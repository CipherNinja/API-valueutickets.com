# GeneratePDF/Flightpdf.py
import openpyxl
from django.http import HttpResponse

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def export_to_excel(modeladmin, request, queryset):
    # Create a workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flight Bookings"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")  # White font for headers
    header_fill = PatternFill(fill_type="solid", start_color="4CAF50")  # Green background
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add column headers with styling
    headers = [
        "Booking ID", "Customer Email", "Passengers", "Payment (Cardholder Name)", "Flight Name",
        "Departure IATA", "Arrival IATA", "Departure Date", "Arrival Date", 
        "Status", "Payable Amount"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Add row data with styling
    for row_num, booking in enumerate(queryset, start=2):  # Start from the second row
        row = [
            booking.booking_id,
            str(booking.customer),  # Customer email
            ", ".join([f"{p.first_name} {p.last_name}" for p in booking.passengers.all()]),
            str(booking.payment.cardholder_name),  # Payment cardholder name
            booking.flight_name,
            booking.departure_iata,
            booking.arrival_iata,
            booking.departure_date.strftime('%Y-%m-%d %H:%M:%S') if booking.departure_date else "",
            booking.arrival_date.strftime('%Y-%m-%d %H:%M:%S') if booking.arrival_date else "",
            booking.status,
            booking.payble_amount,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = center_alignment
            cell.border = thin_border
            if col_num == 10:  # Status column
                # Add conditional color for statuses
                if cell_value == "completed":
                    cell.fill = PatternFill(fill_type="solid", start_color="28A745")  # Green
                elif cell_value == "cancelled":
                    cell.fill = PatternFill(fill_type="solid", start_color="FF5733")  # Red
                elif cell_value == "pending":
                    cell.fill = PatternFill(fill_type="solid", start_color="FFC107")  # Yellow

    # Adjust column widths for better spacing
    column_widths = [15, 25, 30, 25, 20, 15, 15, 20, 20, 15, 18]
    for col_num, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Create HTTP response with Excel content
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename="flight_bookings.xlsx"'
    wb.save(response)
    return response

export_to_excel.short_description = "Export to Excel"
